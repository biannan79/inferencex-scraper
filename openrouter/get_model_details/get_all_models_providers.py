



#!/usr/bin/env python
"""
获取OpenRouter平台所有模型和提供商数据的独立模块

包含功能：
1. 获取所有模型列表 (https://openrouter.ai/api/frontend/models)
2. 获取所有提供商元数据 (https://openrouter.ai/api/frontend/all-providers)

使用Scrapling库进行API请求，解析数据并保存为Excel文件。
"""

import sys
import os
import json
import time
import re
import glob
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
from collections.abc import MutableMapping

# 导入Scrapling库
try:
    from scrapling.fetchers import Fetcher, StealthyFetcher, DynamicFetcher
    SCRAPLING_AVAILABLE = True
except ImportError:
    SCRAPLING_AVAILABLE = False
    print("错误: Scrapling库未安装")
    print("请安装: pip install scrapling[fetchers]")
    print("然后运行: scrapling install")
    # 在except块中定义变量以避免NameError
    Fetcher = None
    StealthyFetcher = None
    DynamicFetcher = None
    sys.exit(1)


# ============================================================================
# API数据获取函数
# ============================================================================

def fetch_api_data(url, description="API数据"):
    """
    使用Scrapling获取API数据

    参数:
        url: API端点URL
        description: 描述信息

    返回:
        dict: 包含success, data, error等信息
    """
    try:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始获取{description}: {url}")

        # 默认请求头
        default_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'Accept-Encoding': 'gzip, deflate',  # 移除br避免Brotli压缩问题
            'Referer': 'https://openrouter.ai/minimax/minimax-m2.7/apps',
            'Origin': 'https://openrouter.ai',
        }

        print(f"  请求头: {default_headers}")
        start_time = time.time()

        # 直接使用StealthyFetcher - 更可靠
        print(f"  使用StealthyFetcher进行API请求...")
        try:
            page = StealthyFetcher.fetch(url, headers=default_headers, headless=True, network_idle=True)
            print(f"  使用StealthyFetcher成功")
        except Exception as fetcher_err:
            print(f"  StealthyFetcher失败: {fetcher_err}")
            print(f"  回退到Fetcher...")
            # 使用Fetcher
            page = Fetcher.get(url, headers=default_headers)
            print(f"  使用Fetcher成功")

        elapsed_time = time.time() - start_time

        print(f"[{datetime.now().strftime('%H:%M:%S')}] Scrapling请求完成，耗时: {elapsed_time:.2f}秒")

        # 调试信息：打印page对象类型和属性
        print(f"  page对象类型: {type(page)}")

        # 打印一些重要属性
        important_attrs = ['body', 'text', 'json', 'content', 'response', 'status_code', 'status', 'code']
        for attr in important_attrs:
            if hasattr(page, attr):
                attr_value = getattr(page, attr)
                print(f"  {attr}: {type(attr_value)}")
                if not callable(attr_value):
                    if attr_value:
                        if isinstance(attr_value, bytes):
                            print(f"    bytes长度: {len(attr_value):,}")
                        else:
                            print(f"    值预览: {str(attr_value)[:100]}...")
                    else:
                        print(f"    值: {attr_value}")

        # 获取响应内容
        raw_text = None
        parsed_data = None

        # 方法1: 尝试直接使用json()方法
        if hasattr(page, 'json') and callable(page.json):
            try:
                parsed_data = page.json()
                print(f"  [方法1] 使用page.json()成功解析JSON数据")
                raw_text = json.dumps(parsed_data, ensure_ascii=False, indent=2)
                print(f"    JSON转换为字符串，大小: {len(raw_text):,} 字符")
            except Exception as json_err:
                print(f"  [方法1] page.json()失败: {json_err}")

        # 方法2: 如果json()失败，尝试body属性
        if not raw_text and hasattr(page, 'body'):
            body_value = page.body
            if body_value:
                print(f"  [方法2] 使用page.body，类型: {type(body_value)}，大小: {len(body_value):,} 字节")
                try:
                    # 如果是bytes，尝试解码
                    if isinstance(body_value, bytes):
                        try:
                            raw_text = body_value.decode('utf-8')
                            print(f"    成功解码为UTF-8字符串")
                        except UnicodeDecodeError:
                            try:
                                raw_text = body_value.decode('latin-1')
                                print(f"    使用latin-1解码")
                            except UnicodeDecodeError:
                                raw_text = str(body_value)
                                print(f"    无法解码，转换为字符串")
                    else:
                        raw_text = str(body_value)

                    # 尝试解析JSON
                    if parsed_data is None:
                        try:
                            parsed_data = json.loads(raw_text)
                            print(f"    成功解析JSON数据")
                        except json.JSONDecodeError as json_err:
                            print(f"    无法解析JSON: {json_err}")
                except Exception as body_err:
                    print(f"  [方法2] 处理page.body时出错: {body_err}")

        # 方法3: 如果上述方法都失败，尝试text属性（可能是TextHandler）
        if not raw_text and hasattr(page, 'text'):
            text_value = page.text
            print(f"  [方法3] 使用page.text，类型: {type(text_value)}")

            # 检查是否是TextHandler对象
            if hasattr(text_value, '__class__') and 'TextHandler' in text_value.__class__.__name__:
                print(f"    text是TextHandler对象")
                # 尝试get()方法
                if hasattr(text_value, 'get') and callable(text_value.get):
                    try:
                        text_result = text_value.get()
                        print(f"    使用get()获取文本，类型: {type(text_result)}")
                        if text_result:
                            raw_text = str(text_result)
                            print(f"      获取到文本，大小: {len(raw_text):,} 字符")
                    except Exception as get_err:
                        print(f"    get()方法失败: {get_err}")

                # 如果get()失败，尝试getall()
                if not raw_text and hasattr(text_value, 'getall') and callable(text_value.getall):
                    try:
                        text_result = text_value.getall()
                        print(f"    使用getall()获取文本，类型: {type(text_result)}")
                        if text_result:
                            raw_text = str(text_result)
                            print(f"      获取到文本，大小: {len(raw_text):,} 字符")
                    except Exception as getall_err:
                        print(f"    getall()方法失败: {getall_err}")
            else:
                # 不是TextHandler，直接使用
                if text_value:
                    raw_text = str(text_value)
                    print(f"    直接使用text值，大小: {len(raw_text):,} 字符")

        # 方法4: 最后尝试，使用str(page)
        if not raw_text:
            raw_text = str(page)
            print(f"  [方法4] 使用str(page)获取内容")

        if not raw_text:
            raw_text = ""
            print("  [警告] 未能获取到响应内容")
        else:
            print(f"  [最终] 响应大小: {len(raw_text):,} 字符")
            if len(raw_text) < 500:
                print(f"  响应内容预览: {raw_text[:500]}")

        # 总是保存原始响应
        save_raw_response(raw_text, url)

        # 返回结果
        if parsed_data is not None:
            return {
                "success": True,
                "data": parsed_data,
                "raw_text": raw_text,
                "elapsed_time": elapsed_time,
                "error": None
            }
        elif raw_text:
            # 有原始文本但没有解析的数据
            # 尝试最后解析JSON（可能在之前的步骤中未尝试）
            try:
                parsed_data = json.loads(raw_text)
                print(f"  成功解析JSON数据")
                return {
                    "success": True,
                    "data": parsed_data,
                    "raw_text": raw_text,
                    "elapsed_time": elapsed_time,
                    "error": None
                }
            except json.JSONDecodeError:
                return {
                    "success": True,
                    "data": raw_text,
                    "raw_text": raw_text,
                    "elapsed_time": elapsed_time,
                    "error": "响应不是JSON格式或无法解析"
                }
        else:
            return {
                "success": False,
                "data": None,
                "raw_text": "",
                "elapsed_time": elapsed_time,
                "error": "未能获取到响应内容"
            }

    except Exception as fetch_err:
        error_msg = f"Scrapling API请求错误: {fetch_err}"
        print(f"[ERROR] {error_msg}")
        import traceback
        traceback.print_exc()

        return {
            "success": False,
            "data": None,
            "error": error_msg
        }


def save_raw_response(raw_text, url, filename=None):
    """
    保存原始API响应到文件

    参数:
        raw_text: 原始响应文本
        url: 请求的URL（用于生成文件名）
        filename: 保存的文件名，如果为None则自动生成
    """
    if filename is None:
        # 从URL生成安全文件名
        parsed_url = urlparse(url)
        path = parsed_url.path.strip('/').replace('/', '_')
        query = parsed_url.query
        if query:
            # 简化查询参数
            query_simple = re.sub(r'[&=]', '_', query)[:50]
            path = f"{path}_{query_simple}"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"raw_response_{path}_{timestamp}.txt"

    output_dir = os.path.dirname(filename) if os.path.dirname(filename) else "."
    os.makedirs(output_dir, exist_ok=True)

    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(raw_text)

        print(f"[OK] 已保存原始响应到: {filename}")
        print(f"     文件大小: {len(raw_text):,} 字符")
        return filename
    except Exception as save_err:
        print(f"[ERROR] 保存原始响应时出错: {save_err}")
        return None


# ============================================================================
# 特定API获取函数
# ============================================================================

def get_all_models():
    """获取所有模型列表"""
    url = "https://openrouter.ai/api/frontend/models"
    result = fetch_api_data(url, "所有模型列表")
    return result


def get_all_providers():
    """获取所有提供商元数据"""
    url = "https://openrouter.ai/api/frontend/all-providers"
    result = fetch_api_data(url, "所有提供商列表")
    return result


def find_latest_model_provider_files():
    """
    查找最新的模型和提供商原始API响应文件
    返回包含文件路径的字典
    """
    patterns = {
        "all_models": "raw_response_api_frontend_models_*.txt",
        "all_providers": "raw_response_api_frontend_all-providers_*.txt"
    }

    latest_files = {}

    for key, pattern in patterns.items():
        files = glob.glob(pattern)
        if files:
            # 按修改时间排序，取最新的文件
            files.sort(key=os.path.getmtime, reverse=True)
            latest_files[key] = files[0]
            print(f"[INFO] 找到最新的{key}文件: {files[0]}")
        else:
            latest_files[key] = None
            print(f"[WARN] 未找到{key}文件，模式: {pattern}")

    return latest_files


# ============================================================================
# 数据清理函数（用于Excel导出）
# ============================================================================

def clean_dataframe_for_excel(df):
    """
    清理DataFrame中的非法控制字符，以便保存到Excel

    参数:
        df: pandas DataFrame

    返回:
        清理后的DataFrame
    """
    import re

    # 定义非法控制字符的正则表达式（ASCII 0x00-0x1F和Unicode U+007F-U+009F）
    # 注意：保留换行符(\n, \r)、制表符(\t)等常见空白字符
    illegal_char_pattern = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

    def clean_value(value):
        """清理单个值中的非法控制字符"""
        if pd.isna(value):
            return value

        # 处理非字符串类型
        if not isinstance(value, str):
            # 对于列表、字典等复杂类型，转换为JSON字符串
            try:
                if isinstance(value, (list, dict, tuple)):
                    return json.dumps(value, ensure_ascii=False)
                else:
                    return str(value)
            except Exception:
                return str(value)

        # 对于字符串，移除非法控制字符
        return illegal_char_pattern.sub('', value)

    # 创建DataFrame的副本
    df_clean = df.copy()

    # 清理所有列
    for col in df_clean.columns:
        try:
            # 使用apply处理每个值，避免数组操作错误
            df_clean[col] = df_clean[col].apply(clean_value)
        except Exception as e:
            print(f"[WARN] 清理列'{col}'时出错: {e}")
            # 如果出错，尝试逐行处理
            for idx in range(len(df_clean)):
                try:
                    original = df_clean.at[idx, col]
                    df_clean.at[idx, col] = clean_value(original)
                except Exception:
                    # 如果仍然出错，设置为空字符串
                    df_clean.at[idx, col] = ""

    return df_clean


def create_excel_file(models_df=None, providers_df=None, output_filename=None, api_info_df=None):
    """
    创建包含模型和提供商数据的Excel文件，可选的API信息工作表作为首页

    参数:
        models_df: 模型DataFrame（可选）
        providers_df: 提供商DataFrame（可选）
        output_filename: 输出文件名（可选，自动生成）
        api_info_df: API信息DataFrame，包含调用的API接口信息（可选）

    返回:
        生成的Excel文件路径
    """
    from openpyxl import Workbook

    if output_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"openrouter_models_providers_{timestamp}.xlsx"

    # 创建Workbook
    wb = Workbook()

    sheets_created = 0

    # 处理API信息工作表（如果提供）
    if api_info_df is not None and not api_info_df.empty:
        try:
            # 清理API信息数据
            api_info_df_clean = clean_dataframe_for_excel(api_info_df)

            # 使用默认工作表作为API信息工作表
            default_sheet = wb.active
            default_sheet.title = "API接口"

            # 写入列名
            for col_idx, col_name in enumerate(api_info_df_clean.columns, start=1):
                default_sheet.cell(row=1, column=col_idx, value=col_name)

            # 写入数据
            for row_idx, row in enumerate(api_info_df_clean.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    default_sheet.cell(row=row_idx, column=col_idx, value=value)

            sheets_created += 1
            print(f"[OK] 已写入API接口信息: {len(api_info_df_clean)}行, {len(api_info_df_clean.columns)}列")
        except Exception as e:
            print(f"[ERROR] 写入API信息工作表时出错: {e}")
            # 如果出错，移除默认工作表
            default_sheet = wb.active
            wb.remove(default_sheet)
    else:
        # 移除默认的工作表（如果没有API信息工作表）
        default_sheet = wb.active
        wb.remove(default_sheet)

    # 写入模型数据
    if models_df is not None and not models_df.empty:
        try:
            # 清理数据
            models_df_clean = clean_dataframe_for_excel(models_df)

            # 创建工作表
            ws_models = wb.create_sheet(title="所有模型")

            # 写入列名
            for col_idx, col_name in enumerate(models_df_clean.columns, start=1):
                ws_models.cell(row=1, column=col_idx, value=col_name)

            # 写入数据
            for row_idx, row in enumerate(models_df_clean.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws_models.cell(row=row_idx, column=col_idx, value=value)

            sheets_created += 1
            print(f"[OK] 已写入模型数据: {len(models_df_clean)}行, {len(models_df_clean.columns)}列")
        except Exception as e:
            print(f"[ERROR] 写入模型数据时出错: {e}")

    # 写入提供商数据
    if providers_df is not None and not providers_df.empty:
        try:
            # 清理数据
            providers_df_clean = clean_dataframe_for_excel(providers_df)

            # 创建工作表
            ws_providers = wb.create_sheet(title="所有提供商")

            # 写入列名
            for col_idx, col_name in enumerate(providers_df_clean.columns, start=1):
                ws_providers.cell(row=1, column=col_idx, value=col_name)

            # 写入数据
            for row_idx, row in enumerate(providers_df_clean.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws_providers.cell(row=row_idx, column=col_idx, value=value)

            sheets_created += 1
            print(f"[OK] 已写入提供商数据: {len(providers_df_clean)}行, {len(providers_df_clean.columns)}列")
        except Exception as e:
            print(f"[ERROR] 写入提供商数据时出错: {e}")

    # 如果没有成功创建工作表，创建一个空的工作表
    if sheets_created == 0:
        ws_empty = wb.create_sheet(title="无数据")
        ws_empty.cell(row=1, column=1, value="未找到可用的模型或提供商数据")
        print("[WARN] 没有数据可写入Excel，创建了空工作表")

    # 保存Excel文件
    try:
        wb.save(output_filename)
        print(f"[SUCCESS] Excel文件已保存: {output_filename}")
        print(f"         文件大小: {os.path.getsize(output_filename):,} 字节")
        return output_filename
    except Exception as e:
        print(f"[ERROR] 保存Excel文件时出错: {e}")
        return None


# ============================================================================
# 主函数
# ============================================================================

def main():
    """主函数：获取所有模型和提供商数据并保存为Excel"""
    print("=" * 60)
    print("OpenRouter 模型和提供商数据获取工具")
    print("=" * 60)

    # 检查Scrapling库是否可用
    if not SCRAPLING_AVAILABLE:
        print("[ERROR] Scrapling库未安装，无法继续")
        sys.exit(1)

    # 获取所有模型数据
    print("\n1. 获取所有模型列表...")
    models_result = get_all_models()

    # 获取所有提供商数据
    print("\n2. 获取所有提供商元数据...")
    providers_result = get_all_providers()

    # 解析数据
    models_df = None
    providers_df = None

    if models_result.get("success") and models_result.get("data"):
        print("\n3. 解析模型数据...")
        try:
            # 保存临时JSON文件用于解析
            temp_file = "temp_models.json"
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(models_result["data"], f, ensure_ascii=False, indent=2)

            models_df = parse_models_to_df(temp_file)
            print(f"   解析成功: {len(models_df)}个模型")

            # 删除临时文件
            os.remove(temp_file)
        except Exception as e:
            print(f"   解析模型数据时出错: {e}")

    if providers_result.get("success") and providers_result.get("data"):
        print("\n4. 解析提供商数据...")
        try:
            # 保存临时JSON文件用于解析
            temp_file = "temp_providers.json"
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(providers_result["data"], f, ensure_ascii=False, indent=2)

            providers_df = parse_frontend_all_providers(temp_file)
            print(f"   解析成功: {len(providers_df)}个提供商")

            # 删除临时文件
            os.remove(temp_file)
        except Exception as e:
            print(f"   解析提供商数据时出错: {e}")

    # 创建API信息DataFrame（作为Excel文件的首页）
    print("\n5. 创建API接口信息工作表...")

    # API端点信息
    api_endpoints = {
        "all_models": "https://openrouter.ai/api/frontend/models",
        "all_providers": "https://openrouter.ai/api/frontend/all-providers"
    }

    # API描述映射
    api_descriptions = {
        "all_models": "获取OpenRouter平台所有模型列表",
        "all_providers": "获取OpenRouter平台所有提供商元数据"
    }

    # 解析函数映射
    parse_function_names = {
        "all_models": "parse_models_to_df",
        "all_providers": "parse_frontend_all_providers"
    }

    # 构建API信息列表
    api_info_rows = []

    # API获取结果
    api_results = {
        "all_models": models_result,
        "all_providers": providers_result
    }

    # 解析结果DataFrame
    api_dataframes = {
        "all_models": models_df,
        "all_providers": providers_df
    }

    for api_name, api_url in api_endpoints.items():
        # 获取API获取状态
        api_result = api_results.get(api_name, {})
        api_fetch_success = api_result.get("success", False)
        fetch_status = "成功" if api_fetch_success else "失败"

        # 获取解析结果
        parse_status = "未解析"
        rows_count = 0
        cols_count = 0
        sheet_name = ""

        # 检查是否有对应的DataFrame
        df = api_dataframes.get(api_name)
        if df is not None and not df.empty:
            parse_status = "成功"
            rows_count = len(df)
            cols_count = len(df.columns)
            sheet_name = "所有模型" if api_name == "all_models" else "所有提供商"
        elif df is not None and df.empty:
            parse_status = "空数据"
        else:
            parse_status = "失败"

        # 获取描述
        description = api_descriptions.get(api_name, "无描述")

        # 获取解析函数名
        parse_function = parse_function_names.get(api_name, "未知")

        api_info_rows.append({
            "API名称": api_name,
            "API URL": api_url,
            "描述": description,
            "获取状态": fetch_status,
            "解析状态": parse_status,
            "解析函数": parse_function,
            "数据行数": rows_count,
            "数据列数": cols_count,
            "生成的工作表": sheet_name
        })

    # 创建API信息DataFrame
    api_info_df = pd.DataFrame(api_info_rows)

    print(f"API接口信息: {len(api_info_df)} 个API")
    for _, row in api_info_df.iterrows():
        print(f"  {row['API名称']}: 获取{row['获取状态']}, 解析{row['解析状态']}, 工作表: {row['生成的工作表']}")

    # 创建Excel文件（包含API信息工作表）
    if models_df is not None or providers_df is not None:
        print("\n6. 创建Excel文件...")
        excel_file = create_excel_file(models_df, providers_df, api_info_df=api_info_df)
        if excel_file:
            print(f"\n[完成] Excel文件已生成: {excel_file}")
        else:
            print("\n[错误] 无法生成Excel文件")
    else:
        print("\n[警告] 没有成功获取到任何数据，无法生成Excel文件")

    print("\n" + "=" * 60)


# ============================================================================
# 解析函数
# ============================================================================

# 展平嵌套字典的辅助函数
def flatten_dict(d, parent_key='', sep='.'):
    """
    递归展平嵌套字典，将嵌套键用点连接。
    处理列表：转换为逗号分隔的字符串。
    """
    from collections.abc import MutableMapping

    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, MutableMapping):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # 将列表转换为字符串，元素用逗号分隔；如果元素是字典则递归处理
            if v and isinstance(v[0], dict):
                # 列表中是对象，则进一步展平（例如 pricing.display_pricing）
                for i, item in enumerate(v):
                    items.extend(flatten_dict(item, f"{new_key}[{i}]", sep=sep).items())
            else:
                # 简单列表，转为字符串
                items.append((new_key, ','.join(map(str, v))))
        else:
            items.append((new_key, v))
    return dict(items)


# 读取 JSON 文件，提取所有模型信息，并返回一个 DataFrame。
def parse_models_to_df(file_path):
    """
    读取 JSON 文件，提取所有模型信息，并返回一个 DataFrame。
    每一行代表一个模型，列名为展平后的字段路径。
    """
    import json
    import pandas as pd

    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    models = data.get('data', [])
    flattened_models = []
    for model in models:
        flat = flatten_dict(model)
        flattened_models.append(flat)

    df = pd.DataFrame(flattened_models)
    return df


# 获取所有模型提供商（如 MiniMax、AtlasCloud 等）的元数据
def parse_frontend_all_providers(file_path: str):
    """解析 /api/frontend/all-providers 接口"""
    import json
    import pandas as pd

    with open(file_path, 'r', encoding='utf-8') as f:
        raw = json.load(f)
    providers = raw.get('data', [])
    rows = []
    for p in providers:
        row = {
            'slug': p.get('slug'),
            'display_name': p.get('displayName'),
            'name': p.get('name'),
            'base_url': p.get('baseUrl'),
            'headquarters': p.get('headquarters'),
            'has_chat_completions': p.get('hasChatCompletions'),
            'has_completions': p.get('hasCompletions'),
            'is_abortable': p.get('isAbortable'),
            'moderation_required': p.get('moderationRequired'),
            'byok_enabled': p.get('byokEnabled'),
            'pricing_strategy': p.get('pricingStrategy'),
            'status_page_url': p.get('statusPageUrl'),
            'icon_url': p.get('icon', {}).get('url') if p.get('icon') else None,
            'data_policy_training': p.get('dataPolicy', {}).get('training'),
            'data_policy_retains_prompts': p.get('dataPolicy', {}).get('retainsPrompts'),
            'owners': p.get('owners'),
            'datacenters': p.get('datacenters')
        }
        rows.append(row)
    return pd.DataFrame(rows)


# ============================================================================
# 主程序入口
# ============================================================================

if __name__ == "__main__":
    main()
